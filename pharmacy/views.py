
from django.shortcuts import render, redirect, get_object_or_404
from .forms import MedicineForm
from .models import Medicine, Bill, BillItem,Supplier, StockHistory
from .forms import MedicineForm, SupplierForm
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from decimal import Decimal
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.utils import timezone
from django.db.models.functions import TruncMonth
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.platypus import Image
import os
from django.conf import settings
from .forms import StoreProfileForm
from .models import Settings
from .forms import SettingsForm
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash



@login_required
def dashboard(request):

    medicine_count = Medicine.objects.count()

    supplier_count = Supplier.objects.count()

    bill_count = Bill.objects.count()

    low_stock = Medicine.objects.filter(
        quantity__lt=10
    ).count()

    low_stock_medicines = Medicine.objects.filter(
    quantity__lt=10
    )

    total_sales = Bill.objects.aggregate(
        Sum('grand_total')
    )['grand_total__sum']

    top_medicines = BillItem.objects.values(
    'medicine__medicine_name'
    ).annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold')[:5]



    today_sales = Bill.objects.aggregate(
    Sum('grand_total')
    )['grand_total__sum'] or 0


    if total_sales is None:
        total_sales = 0

    return render(
        request,
        'dashboard.html',
        {
            'medicine_count': medicine_count,
            'supplier_count': supplier_count,
            'bill_count': bill_count,
            'low_stock': low_stock,
            'total_sales': total_sales,
            'low_stock_medicines': low_stock_medicines,
            'top_medicines': top_medicines,

            'today_sales': today_sales
        }
    )


@login_required
def medicine_list(request):

    search = request.GET.get(
        'search'
    )

    medicines = Medicine.objects.all()

    if search:

        medicines = medicines.filter(
            medicine_name__icontains=search
        )

    return render(
        request,
        'medicine_list.html',
        {
            'medicines': medicines
        }
    )

def add_medicine(request):
    form = MedicineForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('medicine_list')

    return render(request, 'medicine_add.html', {'form': form})

def update_medicine(request, id):
    medicine = get_object_or_404(Medicine, id=id)

    form = MedicineForm(
        request.POST or None,
        instance=medicine
    )

    if form.is_valid():
        form.save()
        return redirect('medicine_list')

    return render(
        request,
        'medicine_add.html',
        {'form': form}
    )

def delete_medicine(request, id):
    medicine = get_object_or_404(
        Medicine,
        id=id
    )

    medicine.delete()

    return redirect('medicine_list')

@login_required
def supplier_list(request):

    suppliers = Supplier.objects.all()

    return render(
        request,
        'supplier_list.html',
        {'suppliers': suppliers}
    )


def add_supplier(request):

    form = SupplierForm(
        request.POST or None
    )

    if form.is_valid():
        form.save()
        return redirect('supplier_list')

    return render(
        request,
        'supplier_add.html',
        {'form': form}
    )


def update_supplier(request, id):

    supplier = get_object_or_404(
        Supplier,
        id=id
    )

    form = SupplierForm(
        request.POST or None,
        instance=supplier
    )

    if form.is_valid():
        form.save()
        return redirect('supplier_list')

    return render(
        request,
        'supplier_add.html',
        {'form': form}
    )


def delete_supplier(request, id):

    supplier = get_object_or_404(
        Supplier,
        id=id
    )

    supplier.delete()

    return redirect('supplier_list')


# def login_page(request):

#     if request.method == 'POST':

#         username = request.POST['username']
#         password = request.POST['password']

#         user = authenticate(
#             request,
#             username=username,
#             password=password
#         )

#         if user is not None:
#             login(request, user)
#             return redirect('dashboard')

#     return render(request, 'login.html')

from django.contrib.auth.models import User

if not User.objects.filter(username="Dipak").exists():
    User.objects.create_superuser(
        username="Dipak",
        password="12345",
        email="dipak@gmail.com"
    )

from django.contrib.auth import authenticate, login
from django.contrib import messages

def login_page(request):

    if request.method == 'POST':

        # username = request.POST.get('username')
        # password = request.POST.get('password')

        # user = authenticate(
        #     request,
        #     username=username,
        #     password=password
        # )
                
        username = request.POST.get('username')
        password = request.POST.get('password')

        print("USERNAME =", username)
        print("PASSWORD =", password)

        user = authenticate(
            request,
            username=username,
            password=password
        )

        print("USER =", user)


        if user is not None:

            login(request, user)

            return redirect('dashboard')

        else:

            messages.error(
                request,
                'Invalid Username or Password'
            )

    return render(
        request,
        'login.html'
    )




from django.contrib import messages

def logout_page(request):

    logout(request)

    messages.success(
        request,
        "You have been logged out successfully."
    )

    return redirect('login')



def invoice_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename="invoice.pdf"'
    )

    p = canvas.Canvas(response)

    p.drawString(
        100,
        750,
        "Medical Store Invoice"
    )

    p.drawString(
        100,
        720,
        "Thank You"
    )

    p.save()

    return response


@login_required
def billing(request):

    medicines = Medicine.objects.all()

    if request.method == "POST":

        customer_name = request.POST.get(
            'customer_name'
        )

        medicine_id = request.POST.get(
            'medicine'
        )

        quantity = int(
            request.POST.get('quantity')
        )

        medicine = Medicine.objects.get(
            id=medicine_id
        )

        # total = medicine.price * quantity
        
        # gst = total * Decimal('0.18')


        # grand_total = total + gst

        total = medicine.price * quantity

        setting = Settings.objects.first()

        if setting:
            gst_percentage = Decimal(str(setting.gst_percentage))
        else:
            gst_percentage = Decimal("18.00")   # Default GST

        gst = (total * gst_percentage) / Decimal("100")

        grand_total = total + gst
        

        quantity = int(
    request.POST.get('quantity')
        )

        if quantity <= 0:

            return render(
                request,
                'billing.html',
                {
                    'medicines': medicines,
                    'error': 'Quantity must be greater than zero'
                }
            )

        if quantity > medicine.quantity:

            return render(
                request,
                'billing.html',
                {
                    'medicines': medicines,
                    'error': 'Not enough stock available'
                }
            )
        
        bill = Bill.objects.create(
            customer_name=customer_name,
            total_amount=total,
            gst_amount=gst,
            grand_total=grand_total
        )

        BillItem.objects.create(
            bill=bill,
            medicine=medicine,
            quantity=quantity,
            price=medicine.price
        )

        medicine.quantity -= quantity
        medicine.save()

        StockHistory.objects.create(
            medicine=medicine,
            sold_quantity=quantity,
            remaining_stock=medicine.quantity
        )



        return redirect('bill_history')

    return render(
        request,
        'billing.html',
        {
            'medicines': medicines
        }
    )

def bill_history(request):

    bills = Bill.objects.all().order_by(
        '-id'
    )

    return render(
        request,
        'bill_history.html',
        {
            'bills': bills
        }
    )

def delete_bill(request, id):

    bill = Bill.objects.get(
        id=id
    )

    bill.delete()

    return redirect(
        'bill_history'
    )

# def invoice(request, id):

#     bill = Bill.objects.get(id=id)

#     items = BillItem.objects.filter(
#         bill=bill
#     )

#     return render(
#         request,
#         'invoice.html',
#         {
#             'bill': bill,
#             'items': items
#         }
# )

def invoice(request, id):

    bill = Bill.objects.get(id=id)

    items = BillItem.objects.filter(bill=bill)

    setting = Settings.objects.first()

    return render(
        request,
        "invoice.html",
        {
            "bill": bill,
            "items": items,
            "setting": setting
        }
    )

def invoice_pdf(request, id):

    bill = Bill.objects.get(id=id)

    items = BillItem.objects.filter(
        bill=bill
    )

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = (
        f'attachment; filename="invoice_{id}.pdf"'
    )

    p = canvas.Canvas(response)

    # OUTER BORDER
    p.rect(20, 40, 550, 760)

    # HEADER
    p.setFillColorRGB(
        0.85,
        0.95,
        1
    )

    p.rect(
        20,
        730,
        550,
        70,
        fill=1
    )

    p.setFillColorRGB(
        0,
        0,
        0
    )

    p.setFont(
        "Helvetica-Bold",
        16
    )

    p.drawCentredString(
        295,
        775,
        "SHINDE MEDICAL STORE"
    )

    p.setFont(
        "Helvetica",
        8
    )

    p.drawString(
        30,
        755,
        "Address : Pune, Maharashtra"
    )

    p.drawString(
        30,
        742,
        "Phone : 9309772996"
    )

    p.drawString(
        30,
        729,
        "GSTIN : 27ABCDE1234F1Z5"
    )

    # INVOICE BOX

    p.rect(
        360,
        700,
        190,
        55
    )

    p.setFont(
        "Helvetica",
        8
    )

    p.drawString(
        370,
        740,
        f"Invoice No : {bill.id}"
    )

    p.drawString(
        370,
        725,
        f"Date : {bill.bill_date.strftime('%d-%m-%Y')}"
    )

    p.drawString(
        370,
        710,
        f"Customer : {bill.customer_name}"
    )

    # GST TITLE

    p.setFont(
        "Helvetica-Bold",
        12
    )

    p.drawCentredString(
        295,
        680,
        "GST INVOICE"
    )

    # TABLE HEADER

    p.setFillColorRGB(
        0.90,
        0.90,
        0.90
    )

    p.rect(
        20,
        640,
        550,
        25,
        fill=1
    )

    p.setFillColorRGB(
        0,
        0,
        0
    )

    # COLUMN LINES

    p.line(60, 640, 60, 180)
    p.line(250, 640, 250, 180)
    p.line(330, 640, 330, 180)
    p.line(430, 640, 430, 180)


    # table_bottom = max(y - 20, 450)

    # p.line(60, 640, 60, table_bottom)
    # p.line(250, 640, 250, table_bottom)
    # p.line(330, 640, 330, table_bottom)
    # p.line(430, 640, 430, table_bottom)

    # HEADER TEXT

    p.setFont(
        "Helvetica-Bold",
        9
    )

    p.drawString(30, 648, "SR")
    p.drawString(80, 648, "MEDICINE")
    p.drawString(270, 648, "QTY")
    p.drawString(350, 648, "PRICE")
    p.drawString(470, 648, "AMOUNT")

    y = 620

    sr = 1

    p.setFont(
        "Helvetica",
        8
    )

    for item in items:

        amount = (
            item.quantity *
            item.price
        )

        p.drawString(
            30,
            y,
            str(sr)
        )

        p.drawString(
            80,
            y,
            item.medicine.medicine_name
        )

   
    p.drawRightString(
            310,
            y,
            str(item.quantity)
        )

    p.drawRightString(
            410,
            y,
            str(item.price)
        )

    p.drawRightString(
            550,
            y,
            str(amount)
        )

    y -= 20
    sr += 1

    # TABLE BOTTOM

    p.line(
        20,
        y,
        570,
        y
    )

    # TOTAL BOX

    total_y = y - 70

    p.rect(
        330,
        total_y,
        220,
        60
    )

    p.setFont(
        "Helvetica",
        9
    )

    p.drawString(
        345,
        total_y + 40,
        f"Subtotal : ₹{bill.total_amount}"
    )

    setting = Settings.objects.first()

    p.drawString(
    345,
    total_y + 25,
    f"GST ({setting.gst_percentage}%) : ₹{bill.gst_amount}"
    )

    p.setFont(
        "Helvetica-Bold",
        10
    )

    p.drawString(
        345,
        total_y + 10,
        f"Grand Total : ₹{bill.grand_total}"
    )

    # FOOTER

    p.line(
        20,
        80,
        570,
        80
    )

    p.setFont(
        "Helvetica",
        8
    )

    p.drawString(
        30,
        60,
        "THANK YOU - VISIT AGAIN"
    )

    p.drawRightString(
        550,
        60,
        "AUTHORIZED SIGNATURE"
    )

    p.save()

    return response


@login_required
def stock_history(request):

    history = StockHistory.objects.all().order_by(
        '-date'
    )

    return render(
        request,
        'stock_history.html',
        {
            'history': history
        }
    )

def home(request):

    if request.user.is_authenticated:

        return redirect('dashboard')

    return redirect('login')

@login_required
def sales_report(request):

    total_sales = Bill.objects.aggregate(
        total=Sum('grand_total')
    )['total'] or 0

    total_bills = Bill.objects.count()

    monthly_sales = Bill.objects.annotate(
        month=TruncMonth('bill_date')
    ).values(
        'month'
    ).annotate(
        total=Sum('grand_total')
    ).order_by('month')

    context = {
        'total_sales': total_sales,
        'total_bills': total_bills,
        'monthly_sales': monthly_sales
    }

    return render(
        request,
        'sales_report.html',
        context
    )

@login_required
def sales_report_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="Sales_Report.pdf"'

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    logo_path = os.path.join(
    settings.BASE_DIR,
    'static',
    'images',
    'logo.png'
)

    if os.path.exists(logo_path):

        logo = Image(
            logo_path,
            width=60,
            height=60
        )

    elements.append(logo)

    title = Paragraph(

        "<font size=20 color='blue'><b>SHINDE MEDICAL STORE</b></font>",

        styles['Title']

    )

    elements.append(title)

    elements.append(

        Paragraph(

            "<b>SALES REPORT</b>",

            styles['Heading2']

        )

    )

    elements.append(Spacer(1,20))

    elements.append(

        Paragraph(

            "Address : Pune, Maharashtra",

            styles['Normal']

        )

    )

    elements.append(

        Paragraph(

            "Phone : 9309772996",

            styles['Normal']

        )

    )

    elements.append(

        Paragraph(

            "GSTIN : 27ABCDE1234F1Z5",

            styles['Normal']

        )

    )

    elements.append(

        Paragraph(

            f"Generated On : {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",

            styles['Normal']

        )

    )

    elements.append(Spacer(1,20))

    total_sales = Bill.objects.aggregate(

        Sum('grand_total')

    )['grand_total__sum'] or 0

    total_bills = Bill.objects.count()

    gst = Bill.objects.aggregate(

        Sum('gst_amount')

    )['gst_amount__sum'] or 0

    avg_bill = 0

    if total_bills > 0:

        avg_bill = round(

            total_sales / total_bills,

            2

        )

    summary = [

        ['Sales Summary',''],

        ['Total Revenue',f'Rs. {total_sales}'],

        ['Total Bills',str(total_bills)],

        ['GST Collected',f'Rs. {gst}'],

        ['Average Bill',f'Rs. {avg_bill}']

    ]

    table = Table(

        summary,

        colWidths=[220,220]

    )

    table.setStyle(

        TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.darkblue),

            ('TEXTCOLOR',(0,0),(-1,0),colors.white),

            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),

            ('GRID',(0,0),(-1,-1),1,colors.black),

            ('BACKGROUND',(0,1),(-1,-1),colors.beige),

            ('BOTTOMPADDING',(0,0),(-1,0),10),

            ('ALIGN',(0,0),(-1,-1),'CENTER')

        ])

    )

    elements.append(table)

    elements.append(Spacer(1,25))

    monthly_sales = Bill.objects.annotate(
    month=TruncMonth('bill_date')
    ).values(
        'month'
    ).annotate(
        total=Sum('grand_total')
    ).order_by('month')

    elements.append(
        Paragraph(
            "<b><font size=14>Month Wise Sales</font></b>",
            styles['Heading2']
        )
    )

    month_data = [
        ['Month', 'Revenue']
    ]

    for sale in monthly_sales:

        month_data.append([
            sale['month'].strftime('%B %Y'),
            f"Rs. {sale['total']}"
        ])

    month_table = Table(
        month_data,
        colWidths=[220,220]
    )

    month_table.setStyle(TableStyle([

        ('BACKGROUND',(0,0),(-1,0),colors.green),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),

        ('GRID',(0,0),(-1,-1),1,colors.black),

        ('BACKGROUND',(0,1),(-1,-1),colors.whitesmoke),

        ('ALIGN',(0,0),(-1,-1),'CENTER')

    ]))

    elements.append(month_table)

    elements.append(Spacer(1,20))    
        

    top_medicines = BillItem.objects.values(
        'medicine__medicine_name'
        ).annotate(
            total_sold=Sum('quantity')
        ).order_by('-total_sold')[:10]

    elements.append(

            Paragraph(

                "<b><font size=14>Top Selling Medicines</font></b>",

                styles['Heading2']

            )

        )

    medicine_data = [

            ['Medicine','Sold Qty']

        ]

    for med in top_medicines:

            medicine_data.append([

                med['medicine__medicine_name'],

                str(med['total_sold'])

            ])

    medicine_table = Table(

            medicine_data,

            colWidths=[220,220]

        )

    medicine_table.setStyle(TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.darkgreen),

            ('TEXTCOLOR',(0,0),(-1,0),colors.white),

            ('GRID',(0,0),(-1,-1),1,colors.black),

            ('BACKGROUND',(0,1),(-1,-1),colors.beige),

            ('ALIGN',(0,0),(-1,-1),'CENTER')

        ]))

    elements.append(medicine_table)

    elements.append(Spacer(1,20))

    low_stock = Medicine.objects.filter(
        quantity__lt=10
    )

    elements.append(

            Paragraph(

                "<b><font size=14>Low Stock Medicines</font></b>",

                styles['Heading2']

            )

        )

    stock_data = [

            ['Medicine','Available Qty']

        ]

    for med in low_stock:

            stock_data.append([

                med.medicine_name,

                str(med.quantity)

            ])

    stock_table = Table(

            stock_data,

            colWidths=[220,220]

        )

    stock_table.setStyle(TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.red),

            ('TEXTCOLOR',(0,0),(-1,0),colors.white),

            ('GRID',(0,0),(-1,-1),1,colors.black),

            ('BACKGROUND',(0,1),(-1,-1),colors.whitesmoke),

            ('ALIGN',(0,0),(-1,-1),'CENTER')

        ]))

    elements.append(stock_table)

    elements.append(Spacer(1,20))


    elements.append(Spacer(1, 20))

    footer = Table(
        [
            ["Generated By", "Medical Store Management System"],
            ["Report Date", datetime.now().strftime("%d-%m-%Y %I:%M %p")]
        ],
        colWidths=[180, 260]
    )

    footer.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0d6efd")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),

        ('BACKGROUND', (0,1), (-1,1), colors.HexColor("#e9ecef")),

        ('GRID', (0,0), (-1,-1), 1, colors.grey),

        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),

        ('ALIGN', (0,0), (-1,-1), 'CENTER'),

        ('BOTTOMPADDING', (0,0), (-1,-1), 8)
    ]))

    elements.append(footer)

    elements.append(Spacer(1, 25))

    elements.append(
        Paragraph(
            "<para align='center'><font size=16 color='green'><b>THANK YOU</b></font></para>",
            styles['Normal']
        )
    )

    elements.append(
        Paragraph(
            "<para align='center'><font size=10>Your business is valuable to us.</font></para>",
            styles['Normal']
        )
    )

    

                    # Generate PDF
    doc.build(elements)

    return response

@login_required
def sales_report_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Sales Report"

    # Header
    ws['A1'] = "SHINDE MEDICAL STORE"
    ws['A2'] = "SALES REPORT"

    ws['A1'].font = Font(
        bold=True,
        size=18
    )

    ws['A2'].font = Font(
        bold=True,
        size=14
    )

    blue_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    # Summary Header
    ws['A4'] = "Summary"

    ws['A4'].fill = blue_fill

    ws['A4'].font = Font(
        bold=True,
        color="FFFFFF"
    )

    total_sales = Bill.objects.aggregate(
        Sum('grand_total')
    )['grand_total__sum'] or 0

    total_bills = Bill.objects.count()

    gst = Bill.objects.aggregate(
        Sum('gst_amount')
    )['gst_amount__sum'] or 0

    ws.append([])

    ws.append(["Total Revenue", total_sales])

    ws.append(["Total Bills", total_bills])

    ws.append(["GST Collected", gst])

    ws.append([])

    # Monthly Sales
    ws.append(["Month", "Revenue"])

    monthly_sales = Bill.objects.annotate(
        month=TruncMonth('bill_date')
    ).values(
        'month'
    ).annotate(
        total=Sum('grand_total')
    )

    for sale in monthly_sales:

        ws.append([
            sale['month'].strftime("%B %Y"),
            float(sale['total'])
        ])

    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="Sales_Report.xlsx"'

    wb.save(response)

    return response

from django.contrib.auth.decorators import login_required
import pandas as pd
import numpy as np

from sklearn.linear_model import LinearRegression

@login_required
def sales_prediction(request):

    medicines = Medicine.objects.all()

    prediction = None
    selected = None
    recommendation = None
    confidence = None
    monthly_data = []

    if request.method == "POST":

        medicine_id = request.POST.get("medicine")

        selected = Medicine.objects.get(id=medicine_id)

        sales = (
            BillItem.objects.filter(medicine=selected)
            .annotate(month=TruncMonth("bill__bill_date"))
            .values("month")
            .annotate(total_qty=Sum("quantity"))
            .order_by("month")
        )

        monthly_data = list(sales)

       
        if len(monthly_data) >= 3:

            x = []
            y = []

            month = 1

            for row in monthly_data:

                x.append([month])
                y.append(row["total_qty"])
                month += 1

            model = LinearRegression()

            model.fit(x, y)

            next_month = [[len(x) + 1]]

            prediction = round(
                model.predict(next_month)[0],
                2
            )

            confidence = "Approx. 90%"

            if prediction > selected.quantity:

                recommendation = (
                    f"⚠ Purchase {int(prediction-selected.quantity)+10} more units."
                )

            else:

                recommendation = "✅ Current stock is sufficient."

        else:

            prediction = "Prediction will be available after at least 3 months of sales history."


    return render(
        request,
        "sales_prediction.html",
        {
            "medicines": medicines,
            "selected": selected,
            "prediction": prediction,
            "recommendation": recommendation,
            "confidence": confidence,
            "monthly_data": monthly_data,
        },
    )


@login_required
def inventory_report(request):

    medicines = Medicine.objects.all()

    total_medicines = medicines.count()

    low_stock = medicines.filter(
        quantity__lt=10
    ).count()

    total_stock = medicines.aggregate(
        total=Sum('quantity')
    )['total'] or 0

    total_value = 0

    for medicine in medicines:
        total_value += medicine.price * medicine.quantity

    return render(
        request,
        'inventory_report.html',
        {
            'medicines': medicines,
            'total_medicines': total_medicines,
            'low_stock': low_stock,
            'total_stock': total_stock,
            'total_value': total_value
        }
    )

@login_required
def inventory_report_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="Inventory_Report.pdf"'

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    logo_path = os.path.join(
        settings.BASE_DIR,
        "static",
        "images",
        "logo.png"
    )

    if os.path.isfile(logo_path):

        logo = Image(
            logo_path,
            width=60,
            height=60
        )

        elements.append(logo)

        elements.append(

        Paragraph(

            "<font size=20 color='blue'><b>SHINDE MEDICAL STORE</b></font>",

            styles['Title']

        )

    )

    elements.append(

        Paragraph(

            "<b>INVENTORY REPORT</b>",

            styles['Heading2']

        )

    )

    elements.append(

        Paragraph(

            f"Generated On : {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",

            styles['Normal']

        )

    )

    elements.append(Spacer(1,20))

    medicines = Medicine.objects.all()

    data = [

        [

            "Medicine",

            "Price",

            "Stock",

            "Total Value"

        ]

    ]

    grand_total = 0

    for med in medicines:

        value = med.price * med.quantity

        grand_total += value

        data.append([

            med.medicine_name,

            f"Rs {med.price}",

            str(med.quantity),

            f"Rs {value}"

        ])

        table = Table(
        data,
        colWidths=[180,80,80,120]
    )

    table.setStyle(

        TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.darkblue),

            ('TEXTCOLOR',(0,0),(-1,0),colors.white),

            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),

            ('GRID',(0,0),(-1,-1),1,colors.black),

            ('BACKGROUND',(0,1),(-1,-1),colors.beige),

            ('ALIGN',(0,0),(-1,-1),'CENTER')

        ])

    )

    elements.append(table)

    elements.append(Spacer(1,20))

    summary = [

        ["Total Medicines", str(medicines.count())],

        ["Total Inventory Value", f"Rs {grand_total}"]

    ]

    summary_table = Table(

        summary,

        colWidths=[220,220]

    )

    summary_table.setStyle(

        TableStyle([

            ('BACKGROUND',(0,0),(-1,-1),colors.lightgrey),

            ('GRID',(0,0),(-1,-1),1,colors.black),

            ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),

            ('ALIGN',(0,0),(-1,-1),'CENTER')

        ])

    )

    elements.append(summary_table)

    elements.append(Spacer(1,20))

    elements.append(

        Paragraph(

            "<b>Generated By :</b> SHINDE MEDICAL STORE",

            styles['Normal']

        )

    )

    doc.build(elements)

    return response

# @login_required
# def profile(request):

#     context = {

#         "store_name": "SHINDE MEDICAL STORE",
#         "owner_name": "Dipak Shinde",
#         "phone": "9309772996",
#         "email": "shindemedical@gmail.com",
#         "gst": "27ABCDE1234F1Z5",
#         "address": "Pune, Maharashtra"

#     }

#     return render(
#         request,
#         "profile.html",
#         context
#     )

from .models import StoreProfile
@login_required
def profile(request):

    profile = StoreProfile.objects.first()

    return render(
        request,
        "profile.html",
        {
            "profile": profile
        }
    )

from .models import *
@login_required
def settings_page(request):

    setting = Settings.objects.first()

    return render(
        request,
        "settings.html",
        {
            "setting": setting
        }
    )

# @login_required
# def edit_profile(request):

#     context = {
#         "store_name": "SHINDE MEDICAL STORE",
#         "owner_name": "Dipak Shinde",
#         "phone": "9309772996",
#         "email": "shindemedical@gmail.com",
#         "gst": "27ABCDE1234F1Z5",
#         "address": "Pune, Maharashtra",
#     }

#     return render(
#         request,
#         "edit_profile.html",
#         context
#     )

@login_required
def edit_profile(request):

    profile = StoreProfile.objects.first()

    if request.method == "POST":

        form = StoreProfileForm(
            request.POST,
            request.FILES,
            instance=profile
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Profile Updated Successfully."
            )

            return redirect("profile")

    else:

        form = StoreProfileForm(instance=profile)

    return render(
        request,
        "edit_profile.html",
        {
            "form": form
        }
    )

# @login_required
# def settings_page(request):

#     setting = Settings.objects.first()

#     if not setting:

#         setting = Settings.objects.create()

#     if request.method == "POST":

#         form = SettingsForm(
#             request.POST,
#             instance=setting
#         )

#         if form.is_valid():

#             form.save()

#             messages.success(
#                 request,
#                 "Settings Updated Successfully."
#             )

#             return redirect("settings")

#     else:

#         form = SettingsForm(
#             instance=setting
#         )

#     return render(
#         request,
#         "settings.html",
#         {
#             "form":form
#         }
#     )

@login_required
def settings_page(request):

    setting = Settings.objects.first()

    if setting is None:
        setting = Settings.objects.create()

    if request.method == "POST":

        form = SettingsForm(
            request.POST,
            instance=setting
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Settings updated successfully."
            )

            return redirect("settings")

    else:

        form = SettingsForm(instance=setting)

    return render(
        request,
        "settings.html",
        {
            "form": form,
            "setting": setting
        }
    )
@login_required
def change_password(request):

    if request.method == "POST":

        form = PasswordChangeForm(
            request.user,
            request.POST
        )

        for field in form.fields.values():
            field.widget.attrs.update({
                "class": "form-control"
            })

        if form.is_valid():

            user = form.save()

            update_session_auth_hash(
                request,
                user
            )

            messages.success(
                request,
                "Password changed successfully."
            )

            return redirect("profile")

    else:

        form = PasswordChangeForm(request.user)

        for field in form.fields.values():
            field.widget.attrs.update({
                "class": "form-control"
            })

    return render(
        request,
        "change_password.html",
        {
            "form": form
        }
    )
